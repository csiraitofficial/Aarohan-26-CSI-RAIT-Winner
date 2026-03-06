import cv2
import numpy as np
import mediapipe as mp
import math
import time
import pickle
import os
from ultralytics import YOLO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


#  CONFIG

VIDEO_SOURCE   = 0                       
FRAME_W        = 854
FRAME_H        = 480
YOLO_IMGSZ     = 320
YOLO_CONF      = 0.38
SKIP_FRAMES    = 2
MAX_TRACK_DIST = 130
MAX_MISSING    = 25
TALKING_DIST   = 220                     
MODEL_PKL      = "xgboost_classroom.pkl" 


#  LOAD XGBOOST MODEL

CLASS_NAMES = ["Attentive", "Distracted", "Looking Down", "Talking", "Using Phone"]
FEATURE_COLS = [
    "elbow_angle_left", "elbow_angle_right",
    "shoulder_angle_left", "shoulder_angle_right",
    "head_pitch", "head_yaw",
    "mouth_open_ratio", "wrist_face_distance",
    "shoulder_distance", "phone_detected",
]

if not os.path.exists(MODEL_PKL):
    raise FileNotFoundError(
        f"\n[ERROR] Model file not found: {MODEL_PKL}\n"
        "Run  train_model.py  first to generate it.\n"
    )

with open(MODEL_PKL, "rb") as f:
    xgb_model = pickle.load(f)
print(f"[INFO] XGBoost model loaded from {MODEL_PKL}")


#  COLOURS  (BGR)

C_GREEN  = (50,  200,  50)
C_RED    = (30,   30, 220)
C_ORANGE = (30,  140, 255)
C_BLUE   = (220, 120,  30)
C_PURPLE = (200,  60, 180)
C_WHITE  = (255, 255, 255)
C_DARK   = (20,   20,  20)
C_PANEL  = (25,   25,  25)
C_CYAN   = (200, 220,  50)

BEHAVIOR_COLORS = {
    "Attentive"   : C_GREEN,
    "Distracted"  : C_ORANGE,
    "Looking Down": C_BLUE,
    "Talking"     : C_PURPLE,
    "Using Phone" : C_RED,
}

BEHAVIOR_SCORES = {
    "Attentive"   : 100,
    "Distracted"  :  75,
    "Looking Down":  80,
    "Talking"     :  70,
    "Using Phone" :  50,
}


#  MEDIAPIPE  –  Face Mesh + Pose (lightweight)

_mp_face = mp.solutions.face_mesh
_face_mesh = _mp_face.FaceMesh(
    static_image_mode=False, max_num_faces=1,
    refine_landmarks=False,
    min_detection_confidence=0.45,
    min_tracking_confidence=0.45,
)

_mp_pose = mp.solutions.pose
_pose    = _mp_pose.Pose(
    static_image_mode=False,
    model_complexity=0,           # lightest model
    smooth_landmarks=True,
    min_detection_confidence=0.45,
    min_tracking_confidence=0.45,
)

# Landmark indices
LM_NOSE   = 1
LM_L_EYE  = 33
LM_R_EYE  = 263
LM_L_MOUTH= 61
LM_R_MOUTH= 291

# Pose landmark indices
PL = _mp_pose.PoseLandmark


def _angle(a, b, c):
    """Angle at vertex b formed by a-b-c (degrees)."""
    ba = np.array(a) - np.array(b)
    bc = np.array(c) - np.array(b)
    cos_a = np.dot(ba, bc) / (np.linalg.norm(ba) * np.linalg.norm(bc) + 1e-6)
    return float(np.degrees(np.arccos(np.clip(cos_a, -1, 1))))


def _dist(a, b):
    return math.hypot(a[0] - b[0], a[1] - b[1])


def extract_features(frame: np.ndarray, box: tuple, phone_flag: int) -> np.ndarray:
    """
    Extract the 10 features from a student crop.
    Returns np.ndarray shape (1, 10) or None if extraction fails.
    """
    x1, y1, x2, y2 = box
    fh, fw = frame.shape[:2]
    x1c, y1c = max(0, x1), max(0, y1)
    x2c, y2c = min(fw, x2), min(fh, y2)

    if x2c <= x1c + 15 or y2c <= y1c + 15:
        return None

    crop = frame[y1c:y2c, x1c:x2c].copy()
    if crop.size == 0:
        return None

    ch, cw = crop.shape[:2]

    # Upscale very small crops
    if cw < 100 or ch < 100:
        scale = max(100 / cw, 100 / ch)
        crop  = cv2.resize(crop, (int(cw * scale), int(ch * scale)),
                           interpolation=cv2.INTER_LINEAR)
        ch, cw = crop.shape[:2]

    rgb = cv2.cvtColor(crop, cv2.COLOR_BGR2RGB)

    
    ela = era = 90.0
    sla = sra = 10.0
    head_pitch = head_yaw = 0.0
    mouth_ratio = 0.05
    wrist_face_dist = 0.65
    shoulder_dist   = 0.35

    # ── POSE – elbow + shoulder angles 
    try:
        pose_res = _pose.process(rgb)
        if pose_res.pose_landmarks:
            lm = pose_res.pose_landmarks.landmark

            def lp(idx):
                p = lm[idx]
                return p.x * cw, p.y * ch

            # Elbow angles
            ela = _angle(lp(PL.LEFT_SHOULDER),  lp(PL.LEFT_ELBOW),  lp(PL.LEFT_WRIST))
            era = _angle(lp(PL.RIGHT_SHOULDER), lp(PL.RIGHT_ELBOW), lp(PL.RIGHT_WRIST))

            # Shoulder tilt (angle between shoulder line and horizontal)
            ls, rs = lp(PL.LEFT_SHOULDER), lp(PL.RIGHT_SHOULDER)
            dx = rs[0] - ls[0]
            dy = rs[1] - ls[1]
            sla = sra = abs(math.degrees(math.atan2(abs(dy), max(abs(dx), 1))))

            # Shoulder width (normalised by crop width)
            shoulder_dist = _dist(ls, rs) / cw

            # Wrist-face distance (use nose from pose if available)
            nose_pose = lp(PL.NOSE)
            lw = lp(PL.LEFT_WRIST)
            rw = lp(PL.RIGHT_WRIST)
            wfd = min(_dist(lw, nose_pose), _dist(rw, nose_pose)) / max(cw, ch)
            wrist_face_dist = min(wfd, 1.0)
    except Exception:
        pass

    # ── FACE MESH – head pitch/yaw + mouth ratio
    try:
        face_res = _face_mesh.process(rgb)
        if face_res.multi_face_landmarks:
            flm = face_res.multi_face_landmarks[0].landmark

            def fp(idx):
                p = flm[idx]
                return p.x * cw, p.y * ch

            nose   = fp(LM_NOSE)
            l_eye  = fp(LM_L_EYE)
            r_eye  = fp(LM_R_EYE)
            l_mth  = fp(LM_L_MOUTH)
            r_mth  = fp(LM_R_MOUTH)

            # Mouth open ratio  
            lip_top    = fp(13)   # upper inner lip
            lip_bottom = fp(14)   # lower inner lip
            face_h     = max(y2c - y1c, 1)
            mouth_ratio = abs(lip_bottom[1] - lip_top[1]) / face_h

            # Yaw (left/right)
            eye_mid_x  = (l_eye[0] + r_eye[0]) / 2.0
            eye_span   = max(abs(r_eye[0] - l_eye[0]), 1.0)
            head_yaw   = float((nose[0] - eye_mid_x) / eye_span * 90)

            # Pitch (up/down)
            eye_mid_y  = (l_eye[1] + r_eye[1]) / 2.0
            head_pitch = float((nose[1] - eye_mid_y) / max(y2c - y1c, 1) * 90)

    except Exception:
        pass

    feat = np.array([[
        ela, era, sla, sra,
        head_pitch, head_yaw,
        mouth_ratio, wrist_face_dist,
        shoulder_dist, float(phone_flag)
    ]], dtype=np.float32)

    return feat


def predict_behavior(feat: np.ndarray) -> tuple:
    """
    Returns (behavior_label_str, confidence_0_to_100, score).
    Falls back to 'Attentive' if prediction fails.
    """
    try:
        pred_idx  = int(xgb_model.predict(feat)[0])
        proba     = xgb_model.predict_proba(feat)[0]
        confidence= int(proba[pred_idx] * 100)
        behavior  = CLASS_NAMES[pred_idx]
        score     = BEHAVIOR_SCORES.get(behavior, 100)
        return behavior, confidence, score
    except Exception:
        return "Attentive", 100, 100



#  TALKING PAIR OVERRIDE

def apply_talking_override(tracks: list):
    """
    If two students are close + facing each other (opposite yaw),
    override both to Talking regardless of XGBoost output.
    """
    for i in range(len(tracks)):
        for j in range(i + 1, len(tracks)):
            a, b = tracks[i], tracks[j]
            ax, ay = a.centre()
            bx, by = b.centre()
            if math.hypot(ax - bx, ay - by) > TALKING_DIST:
                continue
            ya, yb = a.head_yaw, b.head_yaw
            if ya * yb < 0 and abs(ya) > 8 and abs(yb) > 8:
                if a.behavior not in ("Using Phone",):
                    a.behavior = "Talking"
                    a.score    = BEHAVIOR_SCORES["Talking"]
                if b.behavior not in ("Using Phone",):
                    b.behavior = "Talking"
                    b.score    = BEHAVIOR_SCORES["Talking"]



#  STUDENT TRACK

class StudentTrack:
    _next_id = 1

    def __init__(self, raw_box):
        self.id        = StudentTrack._next_id
        StudentTrack._next_id += 1
        b              = np.array(raw_box, dtype=float)
        self._hist     = [b.copy()] * 5
        self._smooth   = b.copy()
        self.missing   = 0
        self.behavior  = "Attentive"
        self.confidence= 100
        self.score     = 100
        self.has_phone = False
        self.head_yaw  = 0.0
        
        # ✅ Store all frame data for Excel
        self.frame_data = []

    def update(self, raw_box):
        b = np.array(raw_box, dtype=float)
        self._hist.append(b)
        if len(self._hist) > 5:
            self._hist.pop(0)
        n = len(self._hist)
        w = np.arange(1, n + 1, dtype=float); w /= w.sum()
        self._smooth = sum(wi * bx for wi, bx in zip(w, self._hist))
        self.missing = 0

    def centre(self):
        x1, y1, x2, y2 = self._smooth
        return (x1 + x2) / 2.0, (y1 + y2) / 2.0

    def ibox(self):
        return tuple(self._smooth.astype(int))
    
    # ✅ Record frame data with all features
    def record_frame(self, features, behavior, confidence, score, phone_flag):
        if features is not None:
            self.frame_data.append({
                'student_id': self.id,
                'elbow_angle_left': float(features[0, 0]),
                'elbow_angle_right': float(features[0, 1]),
                'shoulder_angle_left': float(features[0, 2]),
                'shoulder_angle_right': float(features[0, 3]),
                'head_pitch': float(features[0, 4]),
                'head_yaw': float(features[0, 5]),
                'mouth_open_ratio': float(features[0, 6]),
                'wrist_face_distance': float(features[0, 7]),
                'shoulder_distance': float(features[0, 8]),
                'phone_detected': int(phone_flag),
                'behavior': behavior,
                'confidence': confidence,
                'score': score,
            })



#  TRACKING

def update_tracks(tracks: list, detections: list) -> list:
    used = [False] * len(detections)
    for t in tracks:
        tx, ty = t.centre()
        best_i, best_d = None, MAX_TRACK_DIST
        for i, (x1, y1, x2, y2) in enumerate(detections):
            if used[i]: continue
            d = math.hypot(tx - (x1+x2)/2, ty - (y1+y2)/2)
            if d < best_d: best_d, best_i = d, i
        if best_i is not None:
            t.update(detections[best_i]); used[best_i] = True
        else:
            t.missing += 1
    for i, box in enumerate(detections):
        if not used[i]: tracks.append(StudentTrack(box))
    return [t for t in tracks if t.missing < MAX_MISSING]


def check_phones(tracks: list, phone_boxes: list):
    for t in tracks:
        t.has_phone = False
        tx1, ty1, tx2, ty2 = t.ibox()
        for (px1, py1, px2, py2) in phone_boxes:
            if min(tx2,px2) > max(tx1,px1) and min(ty2,py2) > max(ty1,py1):
                t.has_phone = True; break


#  DRAWING

_FONT = cv2.FONT_HERSHEY_SIMPLEX


def put_label(img, text, x, y, color, fs=0.46, th=1):
    (tw, th_), bl = cv2.getTextSize(text, _FONT, fs, th)
    p = 3
    cv2.rectangle(img, (x-p, y-th_-p), (x+tw+p, y+bl+p), C_DARK, -1)
    cv2.putText(img, text, (x, y), _FONT, fs, color, th, cv2.LINE_AA)


def draw_student(img, t: StudentTrack):
    x1, y1, x2, y2 = t.ibox()
    color = BEHAVIOR_COLORS.get(t.behavior, C_GREEN)

    ov = img.copy()
    cv2.rectangle(ov, (x1-2, y1-2), (x2+2, y2+2), color, 3)
    cv2.addWeighted(ov, 0.35, img, 0.65, 0, img)

    cv2.rectangle(img, (x1, y1), (x2, y2), color, 2)

    L = min(16, (x2-x1)//5, (y2-y1)//5)
    for cx, cy, sx, sy in [(x1,y1,1,1),(x2,y1,-1,1),(x1,y2,1,-1),(x2,y2,-1,-1)]:
        cv2.line(img, (cx, cy), (cx+sx*L, cy),          color, 3)
        cv2.line(img, (cx, cy), (cx,      cy+sy*L),      color, 3)

    put_label(img, f"ID:{t.id}  {t.behavior}", x1, y1-20, color, fs=0.46)
    put_label(img, f"Score:{t.score}  Conf:{t.confidence}%", x1, y1-4, C_WHITE, fs=0.40)


def draw_panel(img, tracks: list, fps: float):
    cnt = {b: 0 for b in CLASS_NAMES}
    for t in tracks:
        if t.behavior in cnt: cnt[t.behavior] += 1

    rows = [
        ("CLASSROOM MONITOR v3",             C_CYAN),
        (f"Total Students  : {len(tracks)}",         C_WHITE),
        (f"Attentive        : {cnt['Attentive']}",    C_GREEN),
        (f"Distracted       : {cnt['Distracted']}",   C_ORANGE),
        (f"Looking Down     : {cnt['Looking Down']}", C_BLUE),
        (f"Talking          : {cnt['Talking']}",      C_PURPLE),
        (f"Using Phone      : {cnt['Using Phone']}",  C_RED),
        (f"FPS              : {fps:.1f}",             C_CYAN),
    ]

    lh, pw, px, py = 22, 235, 8, 8
    ph = len(rows) * lh + 14
    ov = img.copy()
    cv2.rectangle(ov, (px-4, py-4), (px+pw, py+ph), C_PANEL, -1)
    cv2.addWeighted(ov, 0.72, img, 0.28, 0, img)
    for i, (text, color) in enumerate(rows):
        cv2.putText(img, text, (px, py+(i+1)*lh), _FONT, 0.50,
                    color, 2 if i == 0 else 1, cv2.LINE_AA)


# ✅ EXCEL EXPORT - DETAILED FORMAT (Like CSV)
def export_to_excel(tracks: list):
    """Generate Excel with detailed frame-by-frame data"""
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Student Behavior Data"
    
    # Header styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                    top=Side(style='thin'), bottom=Side(style='thin'))
    
    color_map = {
        "Attentive": "C6EFCE",
        "Distracted": "FFEB9C",
        "Looking Down": "BDD7EE",
        "Talking": "E2EFDA",
        "Using Phone": "F4B084"
    }
    
    # Title
    sheet['A1'] = "CLASSROOM BEHAVIOR DETAILED DATA"
    sheet['A1'].font = Font(bold=True, size=13, color="1F4E78")
    sheet.merge_cells('A1:N1')
    sheet['A1'].alignment = Alignment(horizontal="center")
    sheet.row_dimensions[1].height = 22
    
    sheet['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    sheet['A2'].font = Font(italic=True, size=9)
    sheet.merge_cells('A2:N2')
    
    # Column headers
    headers = [
        "Student ID",
        "Elbow Angle L",
        "Elbow Angle R",
        "Shoulder Angle L",
        "Shoulder Angle R",
        "Head Pitch",
        "Head Yaw",
        "Mouth Open Ratio",
        "Wrist Face Distance",
        "Shoulder Distance",
        "Phone Detected",
        "Behavior",
        "Confidence %",
        "Score"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=4, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border
    
    sheet.row_dimensions[4].height = 25
    
    # Collect all frame data
    all_data = []
    for track in tracks:
        all_data.extend(track.frame_data)
    
    # Write data rows
    for row_num, data in enumerate(all_data, 5):
        row_data = [
            data['student_id'],
            f"{data['elbow_angle_left']:.4f}",
            f"{data['elbow_angle_right']:.4f}",
            f"{data['shoulder_angle_left']:.4f}",
            f"{data['shoulder_angle_right']:.4f}",
            f"{data['head_pitch']:.4f}",
            f"{data['head_yaw']:.4f}",
            f"{data['mouth_open_ratio']:.4f}",
            f"{data['wrist_face_distance']:.4f}",
            f"{data['shoulder_distance']:.4f}",
            data['phone_detected'],
            data['behavior'],
            data['confidence'],
            data['score'],
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            
            # Color code behavior column
            if col_num == 12:
                behavior = data['behavior']
                if behavior in color_map:
                    cell.fill = PatternFill(start_color=color_map[behavior], fill_type="solid")
            
            # Number formatting
            if col_num in range(2, 11):
                cell.alignment = Alignment(horizontal="right")
            elif col_num in [13, 14]:
                cell.alignment = Alignment(horizontal="center")
    
    # Column widths
    widths = [12, 13, 13, 15, 15, 12, 10, 16, 18, 16, 14, 16, 13, 8]
    for col_num, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(col_num)].width = width
    
    # Summary sheet
    summary = wb.create_sheet("Summary")
    summary['A1'] = "STUDENT SUMMARY"
    summary['A1'].font = Font(bold=True, size=12, color="1F4E78")
    
    summary['A3'] = "Student ID"
    summary['B3'] = "Total Frames"
    summary['C3'] = "Attentive Count"
    summary['D3'] = "Distracted Count"
    summary['E3'] = "Looking Down Count"
    summary['F3'] = "Talking Count"
    summary['G3'] = "Using Phone Count"
    summary['H3'] = "Avg Confidence"
    summary['I3'] = "Avg Score"
    
    for col in range(1, 10):
        cell = summary.cell(row=3, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    row = 4
    for track in tracks:
        if track.frame_data:
            frames = track.frame_data
            behaviors = [f['behavior'] for f in frames]
            confidences = [f['confidence'] for f in frames]
            scores = [f['score'] for f in frames]
            
            summary[f'A{row}'] = f"Student_{track.id}"
            summary[f'B{row}'] = len(frames)
            summary[f'C{row}'] = behaviors.count("Attentive")
            summary[f'D{row}'] = behaviors.count("Distracted")
            summary[f'E{row}'] = behaviors.count("Looking Down")
            summary[f'F{row}'] = behaviors.count("Talking")
            summary[f'G{row}'] = behaviors.count("Using Phone")
            summary[f'H{row}'] = f"{np.mean(confidences):.1f}"
            summary[f'I{row}'] = f"{np.mean(scores):.1f}"
            row += 1
    
    for col in range(1, 10):
        summary.column_dimensions[get_column_letter(col)].width = 16
    
    filename = f"classroom_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    print(f"\n✅ Excel report saved: {filename}")
    print(f"📊 Total frames recorded: {len(all_data)}")
    print(f"👥 Total students tracked: {len(tracks)}")



#  MAIN LOOP

def main():
    print("=" * 60)
    print("  AI Classroom Monitor  v3.0  –  loading YOLOv8n …")
    model_yolo = YOLO("yolov8n.pt")
    _ = model_yolo(np.zeros((YOLO_IMGSZ, YOLO_IMGSZ, 3), dtype=np.uint8),
                   imgsz=YOLO_IMGSZ, verbose=False)
    print("  YOLOv8n ready.  Q / Esc to quit.")
    print("=" * 60)

    cap = cv2.VideoCapture(VIDEO_SOURCE)
    if not cap.isOpened():
        print(f"[ERROR] Cannot open: {VIDEO_SOURCE}"); return

    cap.set(cv2.CAP_PROP_FRAME_WIDTH,  FRAME_W)
    cap.set(cv2.CAP_PROP_FRAME_HEIGHT, FRAME_H)
    cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)

    WIN = "AI Classroom Monitor v3  [Q = quit]"
    cv2.namedWindow(WIN, cv2.WINDOW_NORMAL)
    cv2.setWindowProperty(WIN, cv2.WND_PROP_FULLSCREEN, cv2.WINDOW_FULLSCREEN)

    tracks:         list[StudentTrack] = []
    cached_persons: list = []
    cached_phones:  list = []
    frame_idx       = 0
    fps_smooth      = 15.0
    t_prev          = time.perf_counter()

    while True:
        ret, frame = cap.read()
        if not ret:
            cap.set(cv2.CAP_PROP_POS_FRAMES, 0)
            ret, frame = cap.read()
            if not ret: break

        frame = cv2.resize(frame, (FRAME_W, FRAME_H))
        frame_idx += 1

        # YOLO detection
        if frame_idx % SKIP_FRAMES == 0:
            results = model_yolo(
                frame, imgsz=YOLO_IMGSZ, conf=YOLO_CONF,
                classes=[0, 67], verbose=False,
            )[0]
            cached_persons.clear(); cached_phones.clear()
            if results.boxes is not None:
                for box in results.boxes:
                    cls  = int(box.cls[0])
                    xyxy = box.xyxy[0].cpu().numpy().astype(int).tolist()
                    (cached_persons if cls == 0 else cached_phones).append(tuple(xyxy))

        # Tracking
        tracks = update_tracks(tracks, cached_persons)
        check_phones(tracks, cached_phones)

        # Feature extraction + prediction
        if frame_idx % SKIP_FRAMES == 0:
            for t in tracks:
                feat = extract_features(frame, t.ibox(), int(t.has_phone))
                if feat is not None:
                    behavior, conf, score = predict_behavior(feat)
                    t.behavior   = behavior
                    t.confidence = conf
                    t.score      = score
                    # ✅ Record frame data
                    t.record_frame(feat, behavior, conf, score, int(t.has_phone))
                    try:
                        t.head_yaw = float(feat[0, 5])
                    except Exception:
                        t.head_yaw = 0.0

            apply_talking_override(tracks)

        # Render
        for t in tracks:
            draw_student(frame, t)

        now        = time.perf_counter()
        fps_smooth = 0.85 * fps_smooth + 0.15 / max(now - t_prev, 1e-6)
        t_prev     = now
        draw_panel(frame, tracks, fps_smooth)

        cv2.imshow(WIN, frame)
        if cv2.waitKey(1) & 0xFF in (ord('q'), ord('Q'), 27):
            break

    cap.release()
    cv2.destroyAllWindows()
    _face_mesh.close()
    _pose.close()
    
    # ✅ Export to Excel
    if tracks:
        export_to_excel(tracks)
    
    print("[INFO] Done.")


if __name__ == "__main__":
    main()